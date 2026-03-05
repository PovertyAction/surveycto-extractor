"""
XLSX Variable Dictionary Exporter
==================================
Reads a variable_dictionary JSON and exports a human-readable Excel workbook.

Usage (via create_variable_dictionaries.py --xlsx):
    exporter = XLSXExporter()
    exporter.export(var_dict_json_path, output_xlsx_path)
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Column definitions: (header, width, wrap_text)
COLUMNS = [
    ("variable_order",                  12,  False),
    ("variable_name",                   28,  False),
    ("stata_type",                      14,  False),
    ("question_text",                   60,  True),
    ("question_type",                   18,  False),
    ("group_path",                      45,  False),
    ("choices_formatted",               50,  True),
    ("choice_list",                     22,  False),
    ("stata_skip_logic",                50,  False),
    ("skip_logic_template",             50,  False),
    ("skip_logic_iteration_specific",   50,  False),
    ("group_relevances",                50,  False),
    ("disabled",                        10,  False),
    ("repeat_iteration",                16,  False),
    ("is_repeat",                       10,  False),
    ("is_select_multiple",              18,  False),
    ("choice_value",                    14,  False),
    ("choice_label",                    30,  False),
    ("is_synthetic",                    12,  False),
    ("repeat_metadata",                 35,  False),
    ("original_variable_name",          28,  False),
]

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")
DISABLED_FILL = PatternFill("solid", fgColor="FFD966")


def _choices_formatted(var_info: dict) -> str:
    """Build the choices_formatted cell value."""
    survey = var_info.get("survey", {})
    choices = survey.get("choices")
    if choices is None:
        return ""

    if isinstance(choices, dict):
        # select_multiple: single choice dict for this binary var
        val   = choices.get("value", "")
        label = choices.get("label", "")
        return f"{val}: {label}"

    if isinstance(choices, list):
        # select_one: pipe-separated list
        parts = []
        for c in choices:
            parts.append(f"{c.get('value', '')}: {c.get('label', '')}")
        joined = " | ".join(parts)
        if len(joined) <= 500:
            return joined
        # Truncate and add count
        truncated = joined[:500]
        shown = sum(1 for p in parts if p in truncated)
        remaining = len(parts) - shown
        return truncated.rstrip(" |") + f" [...{remaining} more]"

    return ""


class XLSXExporter:
    """Export a variable_dictionary JSON to a formatted Excel workbook."""

    def export(self, var_dict_json_path: Path, output_path: Path) -> None:
        var_dict_json_path = Path(var_dict_json_path)
        output_path = Path(output_path)

        print(f"  Loading JSON: {var_dict_json_path}")
        with open(var_dict_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        variables: dict = data.get("variables", {})
        print(f"  {len(variables)} variables to export")

        # Sort by variable_order
        sorted_vars = sorted(
            variables.items(),
            key=lambda kv: kv[1].get("variable_order", 0)
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Variable Dictionary"

        col_names = [c[0] for c in COLUMNS]

        # --- Header row ---
        for col_idx, (col_name, width, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=False)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze pane at B2 (keep variable_order + header visible)
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

        # --- Data rows ---
        for row_idx, (var_name, var_info) in enumerate(sorted_vars, start=2):
            survey  = var_info.get("survey", {})
            is_even = (row_idx % 2 == 0)
            is_disabled = bool(survey.get("disabled", False))

            choices = survey.get("choices")
            choice_value = ""
            choice_label = ""
            if isinstance(choices, dict):
                choice_value = str(choices.get("value", ""))
                choice_label = str(choices.get("label", ""))

            group_relevances_raw = survey.get("group_relevances")
            if isinstance(group_relevances_raw, list):
                group_relevances_str = "; ".join(str(r) for r in group_relevances_raw)
            elif group_relevances_raw is not None:
                group_relevances_str = str(group_relevances_raw)
            else:
                group_relevances_str = ""

            repeat_metadata = var_info.get("repeat_metadata")
            repeat_metadata_str = json.dumps(repeat_metadata) if repeat_metadata else ""

            repeat_iteration = var_info.get("repeat_iteration")
            is_repeat_val    = 1 if repeat_iteration is not None else 0
            is_sm_val        = 1 if isinstance(choices, dict) else 0
            is_synthetic_val = 1 if var_info.get("is_synthetic", False) else 0

            row_values = {
                "variable_order":               var_info.get("variable_order"),
                "variable_name":                var_name,
                "stata_type":                   var_info.get("stata", {}).get("type", ""),
                "question_text":                survey.get("question_text") or "",
                "question_type":                survey.get("type") or "",
                "group_path":                   survey.get("group_path") or "",
                "choices_formatted":            _choices_formatted(var_info),
                "choice_list":                  survey.get("choice_list") or "",
                "stata_skip_logic":             survey.get("stata_skip_logic") or "",
                "skip_logic_template":          survey.get("skip_logic_template") or "",
                "skip_logic_iteration_specific": survey.get("skip_logic_iteration_specific") or "",
                "group_relevances":             group_relevances_str,
                "disabled":                     1 if is_disabled else 0,
                "repeat_iteration":             repeat_iteration if repeat_iteration is not None else "",
                "is_repeat":                    is_repeat_val,
                "is_select_multiple":           is_sm_val,
                "choice_value":                 choice_value,
                "choice_label":                 choice_label,
                "is_synthetic":                 is_synthetic_val,
                "repeat_metadata":              repeat_metadata_str,
                "original_variable_name":       survey.get("original_variable_name") or "",
            }

            for col_idx, (col_name, _, wrap) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_values[col_name])
                cell.alignment = Alignment(wrap_text=wrap, vertical="top")

                if is_disabled:
                    cell.fill = DISABLED_FILL
                elif is_even:
                    cell.fill = ALT_ROW_FILL

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f"  XLSX saved: {output_path}")
